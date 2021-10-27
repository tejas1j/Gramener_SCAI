from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from selenium.common.exceptions import NoSuchElementException
import os
import shutil
import pandas as pd
import csv
from openpyxl import load_workbook
user = ""
pwd = ""
def scrap_data():
        driver = webdriver.Chrome(executable_path='chromedriver.exe')
        driver.get("https://login.microsoftonline.com/e3cf3c98-a978-465f-8254-9d541eeea73c/saml2?SAMLRequest=fdFNT8MwDAbgv1Ll3o%2B0SZpGW6WJXSaNywYcuKA0cVmlJil1ivj5bAMEXHa09PqxbK9Qu7Gc1GaJJ3%2BAtwUwJhtEmOMQ%2FF3wuDiYjzC%2FDwYeD%2Fs1OcU4ocpzM7nMDg48noNWR52Z4BRjVQ6Tyw3mF5kk2zM4eH3RfnvH8Dr4zA1mDhj6GPw4eLj051CZvjKNTHVTy5QJ3qey5CxtLGcUAHRdmatckmS3XZMXWnVSAjc9Z5bKRlAtLOOF7o3tpBFU1IZXpeVSdE3BLYiurDkTRndgKTfszCAusPMYtY9rUha0SQuZFvyBFopyxVkmZPVMkieY8bpFmRUk%2BXCjR%2FV1vTVZZq%2BCxgGV1w5QRaOOm%2Fu9OkfVNIcYTBhJu7rG1XXg%2FFe4Deifd5D25vFX%2BV%2B%2F%2FS7%2F%2F7b9BA%3D%3D&RelayState=https%3A%2F%2Fcpm.dimensiondata.com%2Fepm%2F")
    loginid=""
    while True:
        try:
            loginid = driver.find_element_by_xpath("//*[@id='i0116']")
        except NoSuchElementException as e:
            print(e)
            time.sleep(7)
            driver.refresh()
            time.sleep(7)
        else:
            break
    # loginid = driver.find_element_by_xpath('//*[@id="i0116"]')
    # driver.find_element_by_xpath('//*[@id="idSIButton9"]').click()
    # # password = 
    loginid.send_keys("aswin.s@global.ntt")
    driver.find_element_by_xpath('//*[@id="idSIButton9"]').click()
    time.sleep(5)
    password = driver.find_element_by_xpath("//*[@id='i0118']")
    password.send_keys('V0!dm@in()')
    time.sleep(5)
    driver.find_element_by_xpath("//*[@id='idSIButton9']").click()
    time.sleep(5)
    driver.find_element_by_xpath("//*[@id='idSIButton9']").click()
    while True:
        try:
            driver.find_element_by_xpath("//*[@id='gwt-debug-Results-webChild-0']/div/div[2]/div/div/div/img[3]")
        except NoSuchElementException as e:
            print(e)
            time.sleep(5)
            driver.refresh()
            time.sleep(20)
        else:
            driver.find_element_by_xpath("//*[@id='gwt-debug-Results-webChild-0']/div/div[2]/div/div/div/img[3]").click()
            break
    time.sleep(40)
    driver.save_screenshot("screenshot.png")
    time.sleep(10)
    i=1
    f = open("deed.txt", "r")
    for x in f:
    
        if os.path.exists( os.path.expanduser("~")+'\\Downloads\\Regional Results.xls'):
            os.remove( os.path.expanduser("~")+'\\Downloads\\Regional Results.xls')
        spli=x.split('$')
        spl=spli[1].replace('/','#')
        if os.path.exists((spl.strip('\n')).rstrip()+'.xlsx'):
            os.remove( (spl.strip('\n')).rstrip()+'.xlsx')
        while True:
            try:
                driver.find_element_by_xpath(spli[0])
            except NoSuchElementException as e:
                print(e)
                time.sleep(20)
                driver.refresh()
                time.sleep(20)
            else:
                driver.find_element_by_xpath(spli[0]).click()
                break
        time.sleep(7)

    # //*[@id="gwt-debug-webChild-Regional-Results"]

        while True:
            try:
                driver.find_element_by_xpath("//*[@id='gwt-debug-webChild-Regional-Results']")
            except NoSuchElementException as e:
                print(e)
                time.sleep(15)
                driver.refresh()
                time.sleep(15)
            else:
                driver.find_element_by_xpath("//*[@id='gwt-debug-webChild-Regional-Results']").click()
                break
        time.sleep(5)
        while True:
            try:
                driver.find_element_by_xpath("//*[@id='gwt-debug-menuMainContainer']/fieldset[1]/table/tbody/tr/td/div")
            except NoSuchElementException as e:
                print(e)
                time.sleep(15)
                driver.refresh()
                time.sleep(15)
            else:
                driver.find_element_by_xpath("//*[@id='gwt-debug-menuMainContainer']/fieldset[1]/table/tbody/tr/td/div").click()
                break
        time.sleep(20)
        data=pd.read_excel(os.path.expanduser("~")+'\\Downloads\\Regional Results.xls')
        data.to_excel((spl.rstrip('\n')).rstrip()+'.xlsx')
        data = data.rename(columns=({'Unnamed: 0':'Financial HighLights'}))
        print(data.loc[0])
        if(i==1):
            with open('Revenue.csv', 'w', newline='') as csvFile:
                writer = csv.writer(csvFile)
                writer.writerow(data.columns[0:])
                writer.writerow(data.loc[0])
                i=i+1
        else:
            with open('Revenue.csv', 'a', newline='') as csvFile:
                writer = csv.writer(csvFile)
                writer.writerow(data.loc[0])
        data=pd.read_csv('Revenue.csv',index_col='Financial HighLights')
        data.rename(index = {"Group Services": spl.rstrip()},inplace = True) 
        data.to_csv('Revenue.csv')
        print(data)
        if os.path.exists(os.path.expanduser("~")+'\\Downloads\\Regional Results.xls'):
            os.remove(os.path.expanduser("~")+'\\Downloads\\Regional Results.xls')
        wb = load_workbook(filename = spl.rstrip()+'.xlsx',data_only= True)
        ws=wb.active
        ws['L1'].value="Annotations"
        ws['L2'].value="DD Services Revenue growth lagging behind (+0.04% YoY); 4 out of 5 regions growing while AM (-0.02% YoY) is declining"
        ws['M1'].value="Slide Title"
        ws['M2'].value="DD Services Revenue has been lagging behind for 2Q 2019 when compared with 2Q 2018"
        ws['N1'].value="Values"
        ws['N2'].value="0.04%"
        ws['N3'].value="-0.02%"
        ws['N4'].value="0.01%"
        ws['B1'].value=spl.rstrip().replace(' ','_')
        ws['A1'].value="No"
        wb.save(spl.rstrip()+'.xlsx')
        data=pd.read_excel(spl.rstrip()+'.xlsx')
        print(data.columns)
        data=data[(data[spl.rstrip().replace(' ','_')]!='--- DD AM') & (data[spl.rstrip().replace(' ','_')]!='--- DD AP') & (data[spl.rstrip().replace(' ','_')]!='--- DD AU') & (data[spl.rstrip().replace(' ','_')]!='--- DD EU') & (data[spl.rstrip().replace(' ','_')]!='--- DD MEA')]
        data=data[(data[spl.rstrip().replace(' ','_')]!='--- NTT AM') & (data[spl.rstrip().replace(' ','_')]!='--- NTT AP') & (data[spl.rstrip().replace(' ','_')]!='--- NTT AU') & (data[spl.rstrip().replace(' ','_')]!='--- NTT EU') & (data[spl.rstrip().replace(' ','_')]!='--- NTT MEA')]
        data=data.drop(['No'], axis = 1) 
        data.to_excel(spl.rstrip()+'.xlsx')
        data=pd.read_excel(spl.rstrip()+'.xlsx')
        data[data.columns[2]]=data[data.columns[2]].div(1000000)
        data[data.columns[3]]=data[data.columns[3]].div(1000000)
        data[data.columns[4]]=data[data.columns[4]].div(1000000)
        data[data.columns[5]]=data[data.columns[5]].div(1000000)
        data[data.columns[6]]=data[data.columns[6]].div(1000000)
        data[data.columns[7]]=data[data.columns[7]].mul(100)
        data[data.columns[8]]=data[data.columns[8]].mul(100)
        data[data.columns[9]]=data[data.columns[9]].div(1000000)
        data[data.columns[10]]=data[data.columns[10]].div(1000000)
        data=data.round(1)
        data.to_excel(spl.rstrip()+'.xlsx')
        wb = load_workbook(filename = spl.rstrip()+'.xlsx',data_only= True)
        ws=wb.active
        ws['B1'].value="No"
        wb.save(spl.rstrip()+'.xlsx')
        data=pd.read_excel(spl.rstrip()+'.xlsx')
        data=data.drop(['No'], axis = 1) 
        data.to_excel(spl.rstrip()+'.xlsx')
        driver.back()
    data=pd.read_csv('Revenue.csv')
    data[data.columns[1]]=data[data.columns[1]].div(1000000)
    data[data.columns[2]]=data[data.columns[2]].div(1000000)
    data[data.columns[3]]=data[data.columns[3]].div(1000000)
    data[data.columns[4]]=data[data.columns[4]].div(1000000)
    data[data.columns[5]]=data[data.columns[5]].div(1000000)
    data[data.columns[6]]=data[data.columns[6]].mul(100)
    data[data.columns[7]]=data[data.columns[7]].mul(100)
    data[data.columns[8]]=data[data.columns[8]].div(1000000)
    data[data.columns[9]]=data[data.columns[9]].div(1000000)
    data.round(1)
    data.to_csv('Revenue.csv')
    wb = load_workbook(filename = spl.rstrip()+'.xlsx',data_only= True)
    ws=wb.active
    ws['B1'].value="No"
    wb.save(spl.rstrip()+'.xlsx')
    driver.close()
